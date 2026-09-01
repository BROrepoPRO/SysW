#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Сверка «глубокой подстановки» O/AB против matlib_entries (пункт 4 плана 1.1.3).

Назначение
----------
Бизнес-тест, который сопоставляет фактическое состояние колонок «глубокой
подстановки» на листе `main` рабочей книги work.xlsm с эталоном, хранящимся
в таблице `matlib_entries` базы данных SysW.db.

Что такое «глубокая подстановка O/AB» (по текущей реализации)
-------------------------------------------------------------
При импорте заказ-наряда (ImportFromB2_UI -> ImportDataToMain) при включённом
флаге Mod_Constants.ApplyMatLibSubstitution подставляются модельные артикулы:

  * колонка O (15) — для РАБОТ. Ключ поиска — наименование L(12);
    берётся target_code записи matlib_entries с target_type = 'mod_work'.
    Реализация: Mod_Import.SubstituteWorkArticle.
  * колонка AB (28) — для ЗАПЧАСТЕЙ. Приоритетный ключ — № кат. X(24);
    если X пуст ИЛИ по нему нет совпадения типа 'mod_part' — fallback
    по наименованию Y(25). Берётся target_code записи с target_type='mod_part'.
    Реализация: Mod_Import.SubstitutePartArticle.

Формат matlib_entries (db/schema.sql)
-------------------------------------
  group_name  TEXT   — группа модели (равна main!B14)
  entry_type  TEXT   — 'work' | 'part' (входящая позиция)
  entry_code  TEXT   — поисковый ключ входящей позиции
                      (для работ — наименование in_name листа {G}w;
                       для запчастей — № кат. in_catnum листа {G}z4)
  target_type TEXT   — 'work'|'mod_work'|'part'|'mod_part'
  target_code TEXT   — модельный артикул (значение, которое ДОЛЖНО попасть в O/AB)

Алгоритм сверки
---------------
1. Читаем matlib_entries из SysW.db и строим индекс ожидаемых значений:
       work[norm(entry_code)] -> min(target_code) | target_type='mod_work'
       part[norm(entry_code)] -> min(target_code) | target_type='mod_part'
   (берём первую запись типа в порядке ORDER BY target_type, target_code —
   как FindFirstMatLibIndex в Mod_Import).
2. Читаем лист `main` work.xlsm:
       группа = B14; данные с строки 4.
   Работы — по наличию L(12); ЗЧ — по наличию X(24) или Y(25).
3. Для каждой работы: ожидаемое O = work[norm(L)]; сравниваем с фактич. O(15).
   Для каждой запчасти: ожидаемое AB = part[norm(X)], при отсутствии совпадения
   (или пустом X) — fallback part[norm(Y)]; сравниваем с фактич. AB(28).
4. Нормализация ключей: UPPER + strip (без учёта регистра и краевых пробелов),
   числа приводятся к бездробному виду (19953.0 -> '19953'). Это согласовано с
   провайдером Mod_ModelDBProvider (UCase + Trim).
5. Итог: PASS/FAIL по каждой проверке, перечень расхождений, сводка, exit code.

Обязательный предварительный этап
---------------------------------
Перед запуском сверки должна быть ВЫПОЛНЕНА глубокая подстановка, т.е. импорт
заказ-наряда (макрос ImportFromB2_UI / ImportDataToMain) при включённом
Mod_Constants.ApplyMatLibSubstitution, а группа модели в main!B14 должна
присутствовать в matlib_entries (SysW.db). Иначе O/AB будут пустыми/произвольными
и тест честно покажет расхождения.

Чтение книг выполняется через openpyxl (без COM), БД — через sqlite3 stdlib.
Данные пользователя НЕ модифицируются.
"""
import os
import sys
import time
import sqlite3

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    print("Не удалось импортировать openpyxl:", exc)
    sys.exit(2)

from config import WORKBOOK_PATH, LOGS_DIR, DB_PATH

WORKBOOK = str(WORKBOOK_PATH)
DB_FILE = str(DB_PATH)
REPORT_FILE = str(LOGS_DIR / "oab_reconcile_report.log")

# Колонки листа main (совпадают с Mod_Constants / run_p1_business_test.py)
MAIN_SHEET = "main"
DATA_START_ROW = 4          # данные на main начинаются с 4-й строки
COL_W_NAME = 12             # L  — наименование работы (входящее)
COL_W_MODEL_O = 15          # O  — модельный артикул работы (подстановка)
COL_P_CATNUM = 24           # X  — № кат. (входящее)
COL_P_NAME = 25             # Y  — наименование запчасти (входящее)
COL_P_MODEL_AB = 28         # AB — модельный артикул запчасти (подстановка)
GROUP_CELL = "B14"          # группа модели

# Ожидаемые типы записей matlib_entries
TYPE_MOD_WORK = "mod_work"
TYPE_MOD_PART = "mod_part"


def write_log(message: str):
    """Пишет строку в stdout и в файл отчёта (стиль run_p1_business_test.py)."""
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {message}"
    print(line, flush=True)
    with open(REPORT_FILE, "a", encoding="utf-8") as f:
        f.write(line + "\n")


def norm(v) -> str:
    """Нормализация ключа для сравнения: UPPER + strip, числа без дробной части.

    Согласовано с провайдером Mod_ModelDBProvider (UCase + Trim): сравнение
    идентификаторов/наименований не зависит от регистра и краевых пробелов.
    Целые float приводятся к строке без '.0', чтобы значение 19953.0 совпадало
    с текстовым '19953' (Excel хранит числа как числа, матлаб — как текст).
    """
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v)).strip().upper()
    return str(v).strip().upper()


def val(cell):
    """Безопасное чтение значения ячейки (None -> ''); числа без дробной части."""
    if cell is None:
        return ""
    if isinstance(cell, float) and cell == int(cell):
        return str(int(cell))
    return str(cell).strip()


def load_matlib_index(db_path: str):
    """Читает matlib_entries и строит индексы ожидаемых артикулов по группам.

    Возвращает (groups: set, work_index, part_index), где каждый индекс —
    dict {group: {norm_key: (target_code_count, target_code_first)}}:
      target_code_first — первый артикул типа (мin по ORDER BY target_type,
      target_code), как в FindFirstMatLibIndex; target_code_count — сколько
      совпадений (для диагностики неоднозначности).
    Записи с пустым entry_code исключаются (по ним подстановка невозможна).
    """
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"База данных не найдена: {db_path}")
    conn = sqlite3.connect(db_path)
    try:
        rows = conn.execute(
            "SELECT group_name, entry_type, entry_code, target_type, target_code "
            "FROM matlib_entries"
        ).fetchall()
    finally:
        conn.close()

    groups = set()
    work_index = {}
    part_index = {}
    for group, entry_type, entry_code, target_type, target_code in rows:
        groups.add(group)
        key = norm(entry_code)
        if key == "":
            continue  # нет поискового ключа — подстановка по нему невозможна
        index = work_index if target_type == TYPE_MOD_WORK else (
            part_index if target_type == TYPE_MOD_PART else None)
        if index is None:
            continue
        bucket = index.setdefault(group, {})
        cur = bucket.get(key)
        code = norm(target_code)
        if cur is None:
            bucket[key] = (1, code)
        else:
            cnt, first = cur
            if code < first:  # ORDER BY target_type, target_code -> минимум
                first = code
            bucket[key] = (cnt + 1, first)
    return groups, work_index, part_index


def load_main_rows(workbook_path: str):
    """Читает лист main work.xlsm: группа + списки работ и запчастей.

    Возвращает (group, works, parts):
      group — значение B14 (строка);
      works — список dict {row, in_name, actual_O};
      parts — список dict {row, in_catnum, in_name, actual_AB}.
    """
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if MAIN_SHEET not in wb.sheetnames:
            raise ValueError(f"В {os.path.basename(workbook_path)} нет листа '{MAIN_SHEET}'")
        ws = wb[MAIN_SHEET]
        group = val(ws[GROUP_CELL].value)
        works, parts = [], []
        max_row = ws.max_row
        for r in range(DATA_START_ROW, max_row + 1):
            in_name = val(ws.cell(row=r, column=COL_W_NAME).value)
            model_o = val(ws.cell(row=r, column=COL_W_MODEL_O).value)
            if in_name != "" or model_o != "":
                works.append({"row": r, "in_name": in_name, "actual_O": model_o})
            catnum = val(ws.cell(row=r, column=COL_P_CATNUM).value)
            p_name = val(ws.cell(row=r, column=COL_P_NAME).value)
            model_ab = val(ws.cell(row=r, column=COL_P_MODEL_AB).value)
            if catnum != "" or p_name != "" or model_ab != "":
                parts.append({
                    "row": r, "in_catnum": catnum, "in_name": p_name,
                    "actual_AB": model_ab,
                })
        return group, works, parts
    finally:
        wb.close()


def lookup(index, group, key):
    """Ожидаемый артикул (первый) по группе и ключу; None, если нет."""
    bucket = index.get(group)
    if bucket is None:
        return None
    hit = bucket.get(norm(key))
    return None if hit is None else hit[1]


def reconcile():
    """Выполняет сверку и возвращает код возврата (0 — всё сошлось, 1 — есть FAIL)."""
    os.makedirs(str(LOGS_DIR), exist_ok=True)
    if os.path.exists(REPORT_FILE):
        os.remove(REPORT_FILE)

    write_log("=" * 72)
    write_log("СВЕРКА ГЛУБОКОЙ ПОДСТАНОВКИ O/AB против matlib_entries")
    write_log("=" * 72)

    # --- Шаг 0. Источники ---
    try:
        groups, work_index, part_index = load_matlib_index(DB_FILE)
    except Exception as exc:
        write_log(f"[!] ОШИБКА чтения matlib_entries: {exc}")
        write_log("    Убедитесь, что SysW.db создан (migrate_models_to_sqlite.py).")
        return 2
    write_log(f"[MATLIB] Группы в БД: {sorted(groups)}")

    if not os.path.exists(WORKBOOK):
        write_log(f"[!] ОШИБКА: файл work.xlsm не найден: {WORKBOOK}")
        return 2
    try:
        group, works, parts = load_main_rows(WORKBOOK)
    except Exception as exc:
        write_log(f"[!] ОШИБКА чтения work.xlsm (лист '{MAIN_SHEET}'): {exc}")
        return 2
    write_log(f"[WORKBOOK] {os.path.basename(WORKBOOK)}: группа (B14) = {group!r}, "
              f"строк работ={len(works)}, запчастей={len(parts)}")

    if group == "":
        write_log("[!] ВНИМАНИЕ: группа модели (B14) пуста — сверка невозможна. "
                  "Заполните B14 и повторите.")
        return 1
    if group not in groups:
        write_log(f"[!] ВНИМАНИЕ: группа {group!r} отсутствует в matlib_entries "
                  f"(есть: {sorted(groups)}). Это значит: либо БД устарела "
                  "(нужна миграция/инициация группы), либо подстановка выполнялась "
                  "не по этой БД. Ожидаемые значения будут считаться пустыми.")

    pass_count = 0
    fail_count = 0
    discrepancies = []

    def check(kind, row, detail, actual, expected):
        nonlocal pass_count, fail_count, discrepancies
        ok = norm(actual) == norm(expected)
        if ok:
            pass_count += 1
            write_log(f"[PASS] {kind} row={row} | {detail} | O/AB={actual!r}")
        else:
            fail_count += 1
            discrepancies.append((kind, row, detail, actual, expected))
            write_log(f"[FAIL] {kind} row={row} | {detail} | факт={actual!r} | "
                      f"ожид={expected!r}")

    # --- Работы (O) ---
    write_log("-" * 72)
    write_log("[РАБОТЫ] Сверка колонки O (ключ — наименование L):")
    for w in works:
        expected = lookup(work_index, group, w["in_name"])
        detail = f"L={w['in_name']!r}"
        if expected is None:
            # Совпадения нет -> O должна остаться пустой.
            check("РАБОТА", w["row"], detail, w["actual_O"], "")
        else:
            check("РАБОТА", w["row"], detail, w["actual_O"], expected)

    # --- Запчасти (AB) ---
    write_log("-" * 72)
    write_log("[ЗАПЧАСТИ] Сверка колонки AB (ключ — № кат. X, fallback — Y):")
    for p in parts:
        expected = None
        # 1. Приоритетный ключ — № кат. X(24)
        if p["in_catnum"] != "":
            expected = lookup(part_index, group, p["in_catnum"])
        # 2. Fallback — наименование Y(25), если по X совпадения нет
        if expected is None and p["in_name"] != "":
            expected = lookup(part_index, group, p["in_name"])
        detail = f"X={p['in_catnum']!r} Y={p['in_name']!r}"
        if expected is None:
            check("ЗЧ", p["row"], detail, p["actual_AB"], "")
        else:
            check("ЗЧ", p["row"], detail, p["actual_AB"], expected)

    # --- Сводка ---
    write_log("=" * 72)
    write_log(f"ИТОГ: PASS={pass_count}, FAIL={fail_count}, "
              f"всего проверок={pass_count + fail_count}")
    if discrepancies:
        write_log("ПЕРЕЧЕНЬ РАСХОЖДЕНИЙ:")
        for kind, row, detail, actual, expected in discrepancies:
            write_log(f"  - {kind} row={row} | {detail} | факт={actual!r} | "
                      f"ожид={expected!r}")
    else:
        write_log("Расхождений нет: глубокая подстановка O/AB соответствует matlib_entries.")

    if group not in groups:
        write_log("[!] NB: группа отсутствует в matlib_entries — FAIL обусловлен "
                  "этим, а не некорректной подстановкой.")
    if fail_count > 0:
        write_log("РЕЗУЛЬТАТ: FAIL")
        return 1
    write_log("РЕЗУЛЬТАТ: PASS")
    return 0


def main():
    code = reconcile()
    write_log(f"Завершено, exit code = {code}. Отчёт: {REPORT_FILE}")
    sys.exit(code)


if __name__ == "__main__":
    main()