#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Сборка глобального файла запчастей base/models/z4.xlsx (v1.0.13).

Назначение:
    Формирует единую глобальную базу з/ч (650 000+ позиций), которая
    выносится ИЗ модельных книг в отдельный файл `base/models/z4.xlsx`.

    Источник данных — таблица `parts_catalog` БД `SysW.db` (единый уникальный
    каталог запчастей, заполняемый scripts/migrate_models_to_sqlite.py).
    Если SysW.db отсутствует — скрипт собирает каталог из листа `z4`
    первого модельного файла с таким листом (резервный путь, без БД).

    Структура листа `z4` (совпадает с листом z4 модельной книги):
        заголовки — строка 3, данные — с 4-й строки.
        A=№ п/п, B=Артикул, C=Наименование, D=Ед. изм., E=кол-во,
        F=Цена за ед. изм., G=Кол-во ЗН, H=Сумма ЗН.

Использование:
    python scripts/build_global_parts.py            # сборка из parts_catalog
    python scripts/build_global_parts.py --source models   # резерв: из модельных файлов
"""
from __future__ import annotations

import argparse
import sqlite3
import sys
from pathlib import Path

try:
    from config import (
        DB_PATH,
        MODELS_DIR,
        PROJECT_DIR,
    )
except ImportError:  # pragma: no cover - запуск из другой директории
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from config import (  # type: ignore
        DB_PATH,
        MODELS_DIR,
        PROJECT_DIR,
    )

from openpyxl import Workbook, load_workbook

# Структура листа z4 (заголовки строка 3, данные с 4-й)
DATA_START_ROW = 4
HEADERS = {
    1: "№ п/п",
    2: "Артикул",
    3: "Наименование",
    4: "Ед. изм.",
    5: "кол-во",
    6: "Цена за ед. изм., руб. (в том числе НДС…)",
    7: "Кол-во ЗН",
    8: "Сумма ЗН",
}

OUTPUT_FILE = PROJECT_DIR / "base" / "models" / "z4.xlsx"


def write_catalog(rows) -> None:
    """Записывает каталог [(article, name, unit, price)] в OUTPUT_FILE."""
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "z4"

    for col, text in HEADERS.items():
        ws.cell(row=3, column=col, value=text)

    for idx, (article, name, unit, price) in enumerate(rows, start=1):
        r = DATA_START_ROW + idx - 1
        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=article)
        ws.cell(row=r, column=3, value=name)
        ws.cell(row=r, column=4, value=unit or "")
        ws.cell(row=r, column=5, value="")  # кол-во — не переносится в глобальный каталог
        ws.cell(row=r, column=6, value=price if price else 0)
        ws.cell(row=r, column=7, value="")  # Кол-во ЗН — контекстное, не глобальное
        ws.cell(row=r, column=8, value="")  # Сумма ЗН — формула, не хранится

    wb.save(OUTPUT_FILE)
    print(f"[OK] Глобальная база записана: {OUTPUT_FILE} ({len(rows)} позиций)")


def build_from_db() -> int:
    """Сборка каталога из parts_catalog (SQLite). Возвращает число позиций."""
    if not Path(DB_PATH).exists():
        print(f"[INFO] SysW.db не найден: {DB_PATH}")
        return 0

    conn = sqlite3.connect(DB_PATH)
    try:
        cur = conn.execute(
            "SELECT code, name, unit, price FROM parts_catalog "
            "WHERE code IS NOT NULL AND trim(code) <> '' "
            "ORDER BY code"
        )
        rows = cur.fetchall()
    finally:
        conn.close()

    if rows:
        write_catalog(rows)
    return len(rows)


def build_from_models() -> int:
    """Резервная сборка каталога из листа z4 первого модельного файла."""
    if not MODELS_DIR.exists():
        print(f"[ERROR] Каталог моделей не найден: {MODELS_DIR}")
        return 0

    # Читаем z4 из первого файла с листом z4
    for f in sorted(MODELS_DIR.glob("*.xlsm")) + sorted(MODELS_DIR.glob("*.xlsx")):
        if f.name == "z4.xlsx":
            continue
        try:
            wb = load_workbook(f, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            print(f"[WARN] Не удалось открыть {f.name}: {exc}")
            continue
        if "z4" not in wb.sheetnames:
            wb.close()
            continue
        ws = wb["z4"]
        rows = []
        seen = set()
        for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=8, values_only=True):
            article = str(row[1]).strip() if row[1] is not None else ""
            if not article or article in seen:
                continue
            seen.add(article)
            name = str(row[2]).strip() if row[2] is not None else ""
            unit = str(row[3]).strip() if row[3] is not None else ""
            try:
                price = float(str(row[5]).replace(",", ".")) if row[5] else 0.0
            except ValueError:
                price = 0.0
            rows.append((article, name, unit, price))
        wb.close()
        if rows:
            write_catalog(rows)
            return len(rows)

    print("[WARN] Не найдено ни одного модельного файла с листом z4")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Сборка глобальной базы з/ч z4.xlsx")
    parser.add_argument("--source", choices=["db", "models"], default="db",
                        help="Источник каталога: db (parts_catalog) или models (лист z4)")
    args = parser.parse_args()

    if args.source == "models":
        count = build_from_models()
    else:
        count = build_from_db()
        if count == 0:
            print("[INFO] parts_catalog пуст/отсутствует — резервный сбор из модельных файлов")
            count = build_from_models()

    return 0 if count > 0 else 1


if __name__ == "__main__":
    sys.exit(main())