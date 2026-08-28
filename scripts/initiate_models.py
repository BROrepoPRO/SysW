#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Оркестратор инициации пользовательских модельных файлов в систему SysW (v1.0.15).

Назначение:
    Принимает новые/обновлённые модельные книги групп (base/models/{Group}.xlsm),
    валидирует структуру листов группы, регистрирует группу в таблице model_groups
    и мигрирует данные группы в SysW.db (works, model_works, model_parts,
    matlib_entries, привязки parts к общему каталогу parts_catalog). Отчёт об
    инициации пишется в logs/initiation_report.log.

Сценарии использования:
    python scripts/initiate_models.py            # детекция и миграция новых групп
    python scripts/initiate_models.py --dry-run  # только показать, что будет сделано
    python scripts/initiate_models.py --all      # принудительно мигрировать все группы (идемпотентно)

Замечания:
    - Глобальная база з/ч base/models/z4.xlsx остаётся ОТЛОЖЕННОЙ (R-Z4GEN).
      При появлении файла скрипт лишь сообщает о необходимости сборки через
      scripts/build_global_parts.py, но сам её не запускает.
    - Скрипт НЕ изменяет work.xlsm (критический файл, [E3]): дополнение листов
      models/libname выполняется отдельно после подтверждения пользователя.
      Здесь формируются только рекомендации в отчёте.
"""
from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

# Пути и константы из общей конфигурации
try:
    from config import (
        DB_PATH,
        INITIATION_REPORT_FILE,
        MODELS_DIR,
        DB_SCHEMA_VERSION,
    )
except ImportError:  # pragma: no cover - запуск из другой директории
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from config import (  # type: ignore
        DB_PATH,
        INITIATION_REPORT_FILE,
        MODELS_DIR,
        DB_SCHEMA_VERSION,
    )

from openpyxl import load_workbook

# Схема БД (CREATE TABLE IF NOT EXISTS) и единый источник маппинга
from sqlite_schema import init_db

# Переиспользуемая функция миграции данных группы
from migrate_models_to_sqlite import migrate_group

# Имя служебной глобальной базы з/ч (отложенная задача R-Z4GEN)
GLOBAL_PARTS_FILE = "z4.xlsx"

# Начало данных в листах (заголовки — строка 3, данные — с 4-й)
DATA_START_ROW = 4


def log_line(message: str) -> None:
    """Печать в консоль и добавление строки в лог-файл отчёта."""
    print(message)
    with open(INITIATION_REPORT_FILE, "a", encoding="utf-8") as f:
        f.write(message + "\n")


# ---------------------------------------------------------------------------
# Детекция модельных файлов
# ---------------------------------------------------------------------------

def discover_model_files() -> list:
    """Возвращает отсортированный список модельных файлов групп в base/models/.

    Учитываются *.xlsm и *.xlsx. Исключаются служебные: файлы открытых книг
    (~$*), глобальная база з/ч (z4.xlsx) и прочие не-модельные артефакты.
    """
    files = []
    for pattern in ("*.xlsm", "*.xlsx"):
        for p in MODELS_DIR.glob(pattern):
            if not p.is_file():
                continue
            name = p.name
            if name.startswith("~$"):
                continue
            if name.lower() == GLOBAL_PARTS_FILE:
                continue
            files.append(p)
    return sorted(files, key=lambda p: p.stem)


def registered_groups(conn) -> set:
    """Возвращает множество имён групп, уже зарегистрированных в model_groups."""
    rows = conn.execute("SELECT group_name FROM model_groups").fetchall()
    return {r[0] for r in rows}


# ---------------------------------------------------------------------------
# Валидация структуры листов группы
# ---------------------------------------------------------------------------

def validate_group_structure(wb, group: str) -> list:
    """Проверяет структуру листов модельной книги группы.

    Обязательный лист: {Group} (работы).
    Опциональные: {Group}w (тождества работ), {Group}z4 (тождества з/ч).
    Общий лист z4 (каталог запчастей) допускается, но не обязателен.
    Возвращает список замечаний (пустой — структура корректна).
    """
    issues = []
    names = [s for s in wb.sheetnames]
    lower = {s.lower() for s in names}

    if group.lower() not in lower:
        issues.append(
            f"Отсутствует обязательный лист '{group}' (работы группы)."
        )
    if (group + "w").lower() not in lower:
        issues.append(f"Опциональный лист '{group}w' не найден.")
    if (group + "z4").lower() not in lower:
        issues.append(f"Опциональный лист '{group}z4' не найден.")
    return issues


# ---------------------------------------------------------------------------
# Дополнение work.xlsm (рекомендации, без записи — защита [E3])
# ---------------------------------------------------------------------------

def plan_workbook_update(groups) -> list:
    """Формирует рекомендации по дополнению work.xlsm для новых групп.

    Скрипт не записывает в work.xlsm (критический файл); лист 'models'
    читается в режиме read_only, чтобы указать, какие группы отсутствуют
    в реестре и требуют ручного дополнения.
    """
    recommendations = []
    try:
        from config import WORKBOOK_PATH  # локальный импорт во избежание циклов
        wb = load_workbook(str(WORKBOOK_PATH), read_only=True, data_only=True)
    except Exception as exc:  # файл может быть недоступен или повреждён
        recommendations.append(
            f"work.xlsm недоступен для анализа ({exc}); дополнение листов "
            f"models/libname выполнить вручную."
        )
        return recommendations

    try:
        if "models" in wb.sheetnames:
            ws = wb["models"]
            existing = set()
            for row in ws.iter_rows(min_row=DATA_START_ROW, max_col=2,
                                    values_only=True):
                val = row[1] if len(row) > 1 else None
                if val:
                    existing.add(str(val).strip())
            for g in groups:
                if g not in existing:
                    recommendations.append(
                        f"Лист 'models' work.xlsm: добавить запись группы "
                        f"'{g}' (имя + цена н/ч)."
                    )
        else:
            recommendations.append(
                "В work.xlsm не найден лист 'models' — проверить реестр групп."
            )
    finally:
        wb.close()
    return recommendations


# ---------------------------------------------------------------------------
# Инициация одной группы
# ---------------------------------------------------------------------------

def initiate_group(conn, group: str, src: Path, dry_run: bool = False) -> dict:
    """Валидирует и мигрирует одну новую группу в SysW.db.

    dry_run=True — только валидация и оценка, без записи в БД.
    Возвращает словарь с результатами и списком замечаний.
    """
    result = {"group": group, "file": src.name, "issues": [], "counts": {}}

    try:
        wb = load_workbook(str(src), read_only=True, data_only=True)
    except Exception as exc:
        result["issues"].append(f"Не удалось открыть файл: {exc}")
        return result

    try:
        issues = validate_group_structure(wb, group)
        result["issues"] = issues

        if issues and any("обязательный лист" in msg for msg in issues):
            # Критично: нет листа работ — миграция невозможна.
            return result

        if dry_run:
            # Оценка объёмов без записи: считаем строки работ листа {group}.
            n_works = 0
            if group in wb.sheetnames:
                for row in wb[group].iter_rows(
                    min_row=DATA_START_ROW, max_col=9, values_only=True
                ):
                    if row[1]:
                        n_works += 1
            result["counts"]["works"] = n_works
            return result

        # Миграция данных группы (регистрация в model_groups + заполнение таблиц)
        counts = migrate_group(conn, group, wb, [])
        result["counts"] = counts
        conn.commit()
        return result
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Основной процесс инициации
# ---------------------------------------------------------------------------

def run_initiation(dry_run: bool = False, all_groups: bool = False) -> int:
    """Основной процесс инициации. Возвращает код завершения."""
    start_time = datetime.now()

    files = discover_model_files()
    if not files:
        log_line("  [ОШИБКА] Не найдены модельные файлы в " + str(MODELS_DIR))
        return 1

    # Открываем/создаём схему БД (без полного сброса; drop_tables=False)
    conn = init_db(DB_PATH, drop_tables=False)
    try:
        registered = registered_groups(conn)

        if all_groups:
            # Принудительная инициация всех обнаруженных групп (идемпотентно)
            targets = [p for p in files]
            log_line("  Режим: --all (принудительная инициация всех групп)")
        else:
            # Новые группы: файл есть, в БД нет
            targets = [
                p for p in files
                if p.stem not in registered
            ]
            log_line("  Режим: детекция новых групп")

        if not targets:
            log_line("  Новых групп не обнаружено.")
            log_line("")
            log_line("=" * 60)
            log_line("ОТЧЁТ ОБ ИНИЦИАЦИИ МОДЕЛЬНЫХ ФАЙЛОВ")
            log_line("=" * 60)
            log_line(f"Дата: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
            log_line(f"Зарегистрировано групп в SysW.db: {len(registered)}")
            log_line(f"Режим: {'--dry-run' if dry_run else 'применение'}")
            log_line("Новых групп: нет")
            return 0

        # Инициализация отчёта (перезаписываем при каждом прогоне)
        with open(INITIATION_REPORT_FILE, "w", encoding="utf-8") as f:
            f.write("")

        log_line("=" * 60)
        log_line("ОТЧЁТ ОБ ИНИЦИАЦИИ МОДЕЛЬНЫХ ФАЙЛОВ")
        log_line("=" * 60)
        log_line(f"Дата: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
        log_line(f"База данных: {DB_PATH}")
        log_line(f"Режим: {'--dry-run (только оценка)' if dry_run else 'применение'}")
        log_line("")

        # Проверка появления глобальной базы з/ч (отложенная R-Z4GEN)
        global_parts = MODELS_DIR / GLOBAL_PARTS_FILE
        if global_parts.exists():
            log_line(
                f"  [ИНФО] Обнаружена глобальная база з/ч {GLOBAL_PARTS_FILE}. "
                f"Для сборки каталога требуется запуск "
                f"scripts/build_global_parts.py (задача R-Z4GEN)."
            )

        total = {
            "works": 0,
            "model_works": 0,
            "model_parts": 0,
            "matlib_entries": 0,
        }
        initiated = []
        for p in targets:
            log_line(f"  Группа '{p.stem}' ({p.name})...")
            res = initiate_group(conn, p.stem, p, dry_run=dry_run)
            if res["issues"]:
                for msg in res["issues"]:
                    log_line(f"    [ЗАМЕЧАНИЕ] {msg}")
            if res["counts"]:
                total["works"] += res["counts"].get("works", 0)
                total["model_works"] += res["counts"].get("model_works", 0)
                total["model_parts"] += res["counts"].get("model_parts", 0)
                total["matlib_entries"] += res["counts"].get("matlib_entries", 0)
                log_line(
                    f"    works={res['counts'].get('works', 0)}, "
                    f"model_works={res['counts'].get('model_works', 0)}, "
                    f"model_parts={res['counts'].get('model_parts', 0)}"
                )
                initiated.append(res["group"])

        # Рекомендации по дополнению work.xlsm
        if initiated and not dry_run:
            log_line("")
            log_line("  Рекомендации по дополнению work.xlsm (листы models/libname):")
            for rec in plan_workbook_update(initiated):
                log_line(f"    - {rec}")

        # Итоговые сведения
        log_line("")
        log_line(f"Обработано групп: {len(targets)}")
        if initiated:
            log_line(f"Инициировано групп: {len(initiated)}")
            log_line(f"Всего строк works (оценка/вставка): {total['works']}")
            log_line(f"Всего тождеств работ (model_works): {total['model_works']}")
            log_line(f"Всего тождеств запчастей (model_parts): {total['model_parts']}")
            log_line(f"Всего записей matlib_entries: {total['matlib_entries']}")
        dur = (datetime.now() - start_time).total_seconds()
        log_line(f"Время выполнения: {dur:.2f} с")
        log_line("=" * 60)
        return 0
    finally:
        conn.close()


def main() -> int:
    """Точка входа CLI."""
    parser = argparse.ArgumentParser(
        description="Инициация пользовательских модельных файлов в SysW.db"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Только показать, какие группы будут инициированы, без записи в БД",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Принудительно инициировать все обнаруженные группы (идемпотентно)",
    )
    args = parser.parse_args()

    return run_initiation(dry_run=args.dry_run, all_groups=args.all)


if __name__ == "__main__":
    sys.exit(main())