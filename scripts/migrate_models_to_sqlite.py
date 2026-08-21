#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Конвертер модельных файлов base/models/*.xlsm в единую базу SysW.db (SQLite).

Назначение:
    Переносит данные 6 модельных групп (4x4, 2170, 2180, 2190, GAZ, UAZ)
    из Excel-файлов (openpyxl, БЕЗ COM/Excel) в таблицы SysW.db:
    model_groups, works, parts, model_works, model_parts, matlib_entries.

Ключевые свойства:
    - Только stdlib + openpyxl (без win32com — обходит COM-зависание).
    - load_workbook(read_only=True, data_only=True) — читает ЗНАЧЕНИЯ (кэшированные),
      не формулы. None (формулы без кэша) -> 0 / пустая строка.
    - Идемпотентность: повторный запуск не дублирует данные (DELETE + INSERT
      в транзакции; пересоздание при --force).
    - Производительность: потоковое чтение с остановкой при длинной серии пустых
      строк (огромные форматированные листы z4/{}z4 не сканируются до конца);
      массовые вставки через executemany; коммит после каждой группы.
    - Лист z4 (все запчасти) ОБЩИЙ для всех групп: читается ОДИН раз из первого
      файла с листом z4 и вставляется для КАЖДОЙ группы.
    - Перед перезаписью существующего SysW.db создаётся резервная копия
      в _backup/SysW_<дата>.db.
    - PRAGMA user_version = 1, journal_mode = WAL (конкурентный доступ Python+VBA).
    - Отчёт о миграции пишется в logs/migration_report.log (путь из config.py).

Использование:
    python scripts/migrate_models_to_sqlite.py            # обычный прогон
    python scripts/migrate_models_to_sqlite.py --force    # полный пересозданный прогон

Маппинг листов -> таблицы (заголовки строка 3, данные с 4-й строки):
    {GroupName}   -> works         (B->code, C->name, D->norm_hours, F->price)
    z4 (один раз) -> parts         (B->code, C->name, D->unit, F->price)
    {GroupName}w  -> model_works   (B->out_article, C->out_name, D->norm_hours,
                                    G->qty_zn, I->aggregate, J->in_name)
    {GroupName}z4 -> model_parts   (B->out_article, C->out_name, G->qty_zn,
                                    F->price, I->aggregate, J->in_catnum, K->in_name)
    служебные {NN}M -> не переносятся.

Критерий строки тождества (совпадает с VBA-логикой Mod_ModelDB):
    не пустой B (OutArticle) И не пустой I (Агрегат).

Примечание о маппинге: фактическая структура листов моделей (по docs/table.md,
раздел 0.7) отличается от условной таблицы плана (A->code): у всех 6 групп
код находится в столбце B (Артикул), наименование — в C. Поэтому конвертер
использует фактические колонки (образец пользователя), а не план 2.3.
"""
from __future__ import annotations

import argparse
import shutil
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

# Пути и константы из общей конфигурации
try:
    from config import (
        DB_PATH,
        MIGRATION_REPORT_FILE,
        MODELS_DIR,
        PROJECT_DIR,
        DB_SCHEMA_VERSION,
    )
except ImportError:  # pragma: no cover - запуск из другой директории
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from config import (  # type: ignore
        DB_PATH,
        MIGRATION_REPORT_FILE,
        MODELS_DIR,
        PROJECT_DIR,
        DB_SCHEMA_VERSION,
    )

from openpyxl import load_workbook

# Структура схемы — единый источник (scripts/sqlite_schema.py)
from sqlite_schema import init_db, set_user_version

# Версия схемы для PRAGMA user_version
SCHEMA_VERSION = DB_SCHEMA_VERSION

# Начало данных в листах (заголовки — строка 3, данные — с 4-й)
DATA_START_ROW = 4

# Допуск: подряд идущих полностью пустых строк после появления данных,
# после которого чтение листа прекращается (защита от гигантских
# форматированных листов z4/{}z4).
EMPTY_TOLERANCE = 100


# ---------------------------------------------------------------------------
# Вспомогательные функции
# ---------------------------------------------------------------------------

def cell_text(value) -> str:
    """Приводит значение ячейки к строке; None -> пустая строка."""
    if value is None:
        return ""
    return str(value).strip()


def cell_num(value) -> float:
    """Приводит числовое значение ячейки; None -> 0.0; строки с запятой -> число."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s == "":
        return 0.0
    # Поддержка десятичного разделителя-запятой (русская локаль Excel)
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def is_row_empty(row) -> bool:
    """Возвращает True, если все значения строки пусты/None."""
    for v in row:
        if v is not None and str(v).strip() != "":
            return False
    return True


def data_rows(ws, max_col: int, start_row: int = DATA_START_ROW):
    """Генерирует пары (номер_строки, значения_строки) для строк с данными.

    Чтение потоковое: останавливается при встрече длинной серии полностью
    пустых строк ПОСЛЕ появления хотя бы одной непустой строки. Это позволяет
    не сканировать до конца огромные форматированные листы.
    """
    empty_run = 0
    seen_data = False
    for i, row in enumerate(
        ws.iter_rows(min_row=start_row, max_col=max_col, values_only=True),
        start=start_row,
    ):
        if not is_row_empty(row):
            empty_run = 0
            seen_data = True
            yield i, row
        elif seen_data:
            empty_run += 1
            if empty_run >= EMPTY_TOLERANCE:
                break


def log_line(message: str) -> None:
    """Печать в консоль и добавление строки в лог-файл отчёта."""
    print(message)
    with open(MIGRATION_REPORT_FILE, "a", encoding="utf-8") as f:
        f.write(message + "\n")


# ---------------------------------------------------------------------------
# Извлечение данных из листов одной группы
# ---------------------------------------------------------------------------

def extract_works(conn: sqlite3.Connection, group: str, wb) -> int:
    """Заполняет таблицу works из листа {group}. Возвращает число строк."""
    sheet = group
    if sheet not in wb.sheetnames:
        return 0
    ws = wb[sheet]
    rows = []
    for _, row in data_rows(ws, max_col=9):  # A..I
        code = cell_text(row[1])       # B: Артикул
        if not code:
            continue
        rows.append(
            (group, code, cell_text(row[2]), "", cell_num(row[3]),
             cell_num(row[5]), "")
        )
    if rows:
        conn.executemany(
            "INSERT OR REPLACE INTO works(group_name, code, name, unit, "
            "norm_hours, price, note) VALUES (?, ?, ?, ?, ?, ?, ?)",
            rows,
        )
    return len(rows)


def extract_parts(conn: sqlite3.Connection, group: str, wb) -> int:
    """Заполняет таблицу parts из листа z4 группы. Возвращает число строк."""
    sheet = "z4"
    if sheet not in wb.sheetnames:
        return 0
    ws = wb[sheet]
    rows = []
    for _, row in data_rows(ws, max_col=8):  # A..H
        code = cell_text(row[1])       # B: Артикул
        if not code:
            continue
        rows.append(
            (group, code, cell_text(row[2]), cell_text(row[3]),
             cell_num(row[5]), "")
        )
    if rows:
        conn.executemany(
            "INSERT OR REPLACE INTO parts(group_name, code, name, unit, "
            "price, note) VALUES (?, ?, ?, ?, ?, ?)",
            rows,
        )
    return len(rows)


def extract_model_works(conn: sqlite3.Connection, group: str, wb) -> int:
    """Заполняет model_works и matlib_entries из листа {group}w.

    Возвращает число тождеств работ (строк model_works).
    """
    sheet = group + "w"
    if sheet not in wb.sheetnames:
        return 0
    ws = wb[sheet]
    mw_rows = []
    ml_rows = []
    for _, row in data_rows(ws, max_col=13):  # A..M
        out_article = cell_text(row[1])   # B
        aggregate = cell_text(row[8])     # I
        if not out_article or not aggregate:
            continue
        out_name = cell_text(row[2])      # C
        norm_hours = cell_num(row[3])     # D
        qty_zn = cell_num(row[6])         # G
        in_name = cell_text(row[9])       # J
        mw_rows.append(
            (group, out_article, out_name, norm_hours, qty_zn, aggregate,
             in_name, "")
        )
        ml_rows.append((group, in_name, out_name, out_article, out_name, qty_zn))
    if mw_rows:
        conn.executemany(
            "INSERT OR REPLACE INTO model_works(group_name, out_article, "
            "out_name, norm_hours, qty_zn, aggregate, in_name, note) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            mw_rows,
        )
        conn.executemany(
            "INSERT INTO matlib_entries(group_name, entry_type, entry_code, "
            "entry_name, target_type, target_code, target_name, coefficient) "
            "VALUES (?, 'work', ?, ?, 'mod_work', ?, ?, ?)",
            ml_rows,
        )
    return len(mw_rows)


def extract_model_parts(conn: sqlite3.Connection, group: str, wb) -> int:
    """Заполняет model_parts и matlib_entries из листа {group}z4.

    Возвращает число тождеств запчастей (строк model_parts).
    """
    sheet = group + "z4"
    if sheet not in wb.sheetnames:
        return 0
    ws = wb[sheet]
    mp_rows = []
    ml_rows = []
    for _, row in data_rows(ws, max_col=14):  # A..N
        out_article = cell_text(row[1])   # B
        aggregate = cell_text(row[8])     # I
        if not out_article or not aggregate:
            continue
        out_name = cell_text(row[2])      # C
        qty_zn = cell_num(row[6])         # G
        price = cell_num(row[5])          # F
        in_catnum = cell_text(row[9])     # J: № кат.
        in_name = cell_text(row[10])      # K: Наим-ние
        mp_rows.append(
            (group, out_article, out_name, qty_zn, price, aggregate,
             in_catnum, in_name, "")
        )
        ml_rows.append((group, in_catnum, in_name, out_article, out_name, qty_zn))
    if mp_rows:
        conn.executemany(
            "INSERT OR REPLACE INTO model_parts(group_name, out_article, "
            "out_name, qty_zn, price, aggregate, in_catnum, in_name, note) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
            mp_rows,
        )
        conn.executemany(
            "INSERT INTO matlib_entries(group_name, entry_type, entry_code, "
            "entry_name, target_type, target_code, target_name, coefficient) "
            "VALUES (?, 'part', ?, ?, 'mod_part', ?, ?, ?)",
            ml_rows,
        )
    return len(mp_rows)


def load_shared_parts(wb) -> list:
    """Читает лист z4 ОДИН раз (общий каталог запчастей для всех групп).

    Возвращает список кортежей (code, name, unit, price) из листа z4.
    """
    if "z4" not in wb.sheetnames:
        return []
    ws = wb["z4"]
    out = []
    for _, row in data_rows(ws, max_col=8):  # A..H
        code = cell_text(row[1])       # B: Артикул
        if not code:
            continue
        out.append(
            (code, cell_text(row[2]), cell_text(row[3]), cell_num(row[5]))
        )
    return out


def insert_shared_parts_for_group(conn: sqlite3.Connection, group: str,
                                  shared: list) -> None:
    """Вставляет общий каталог запчастей для одной группы."""
    if not shared:
        return
    rows = [(group, code, name, unit, price, "") for code, name, unit, price in shared]
    conn.executemany(
        "INSERT OR REPLACE INTO parts(group_name, code, name, unit, price, note) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        rows,
    )


# ---------------------------------------------------------------------------
# Валидация результатов
# ---------------------------------------------------------------------------

def validate_db(conn: sqlite3.Connection) -> None:
    """Выполняет PRAGMA integrity_check и логирует результат."""
    row = conn.execute("PRAGMA integrity_check").fetchone()
    status = row[0] if row else "unknown"
    log_line(f"  PRAGMA integrity_check: {status}")
    if status != "ok":
        log_line("  [ОШИБКА] Целостность базы данных нарушена!")


def run_migration(force: bool = False) -> int:
    """Основной процесс миграции. Возвращает код завершения."""
    # Определяем группы моделей сканированием каталога base/models/*.xlsm
    groups = sorted(
        p.stem for p in MODELS_DIR.glob("*.xlsm")
        if p.is_file()
    )
    if not groups:
        log_line("  [ОШИБКА] Не найдены модельные файлы *.xlsm в " + str(MODELS_DIR))
        return 1

    # Логирование времени начала
    start_time = datetime.now()

    # Бэкап существующего SysW.db перед перезаписью
    if DB_PATH.exists():
        backup_dir = PROJECT_DIR / "_backup"
        backup_dir.mkdir(exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"SysW_{stamp}.db"
        try:
            shutil.copy2(str(DB_PATH), str(backup_path))
            log_line(f"  Резервная копия SysW.db -> {backup_path.name}")
        except OSError as exc:
            log_line(f"  [ПРЕДУПРЕЖДЕНИЕ] Не удалось создать бэкап: {exc}")

    # Открываем/создаём схему БД (drop_tables = force => полное пересоздание)
    conn = init_db(DB_PATH, drop_tables=force)

    # DELETE + INSERT в транзакции (идемпотентность при обычном прогоне)
    if not force:
        tables = [
            "matlib_entries",
            "model_works",
            "model_parts",
            "works",
            "parts",
            "model_groups",
        ]
        for t in tables:
            conn.execute('DELETE FROM "{}";'.format(t))

    try:
        # --- Загрузка общего каталога запчастей z4 (один раз) ---
        shared_parts = []
        parts_read_once = False
        for g in groups:
            src = MODELS_DIR / f"{g}.xlsm"
            if not src.exists():
                continue
            wb = load_workbook(str(src), read_only=True, data_only=True)
            if "z4" in wb.sheetnames and not parts_read_once:
                shared_parts = load_shared_parts(wb)
                parts_read_once = True
                log_line(f"  Лист z4 прочитан ОДИН раз из файла '{g}.xlsm': "
                         f"{len(shared_parts)} запчастей (общий каталог)")
                wb.close()
                break
            wb.close()
        if not parts_read_once:
            log_line("  [ПРЕДУПРЕЖДЕНИЕ] Лист z4 не найден ни в одном файле группы")

        total = {
            "works": 0,
            "parts_inserted": 0,      # всего вставлено в parts (по группам)
            "parts_shared": len(shared_parts),  # уникальный каталог z4
            "model_works": 0,
            "model_parts": 0,
            "matlib_entries": 0,
        }
        for g in groups:
            src = MODELS_DIR / f"{g}.xlsm"
            if not src.exists():
                log_line(f"  [ПРОПУСК] Файл группы '{g}' не найден: {src}")
                continue
            log_line(f"  Обработка группы '{g}' ({src.name})...")
            wb = load_workbook(str(src), read_only=True, data_only=True)

            # Регистрация группы (идемпотентно)
            conn.execute(
                "INSERT OR REPLACE INTO model_groups(group_name) VALUES (?)", (g,)
            )

            n_w = extract_works(conn, g, wb)
            n_mw = extract_model_works(conn, g, wb)
            n_mp = extract_model_parts(conn, g, wb)
            # Общий каталог запчастей вставляется для КАЖДОЙ группы
            insert_shared_parts_for_group(conn, g, shared_parts)

            total["works"] += n_w
            total["model_works"] += n_mw
            total["model_parts"] += n_mp
            total["matlib_entries"] += n_mw + n_mp
            total["parts_inserted"] += len(shared_parts)

            log_line(f"    works={n_w}, model_works={n_mw}, "
                     f"model_parts={n_mp}, parts={len(shared_parts)}")
            wb.close()
            conn.commit()

        # Выставляем версию схемы и WAL-режим
        set_user_version(conn, SCHEMA_VERSION)
        conn.execute("PRAGMA journal_mode=WAL")
        conn.commit()

        # Валидация целостности и подсчёт строк в БД
        db_counts = {}
        for tbl in ["works", "parts", "model_works", "model_parts",
                    "matlib_entries", "model_groups"]:
            n = conn.execute(f'SELECT COUNT(*) FROM "{tbl}"').fetchone()[0]
            db_counts[tbl] = n
        validate_db(conn)

    finally:
        conn.close()

    # ------------------------------------------------------------------
    # Формирование отчёта
    # ------------------------------------------------------------------
    log_line("")
    log_line("=" * 60)
    log_line("ОТЧЁТ О МИГРАЦИИ МОДЕЛЕЙ В SysW.db")
    log_line("=" * 60)
    log_line(f"Дата: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    log_line(f"Группы ({len(groups)}): {', '.join(groups)}")
    log_line(f"База данных: {DB_PATH}")
    log_line(f"Режим: {'--force (полное пересоздание)' if force else 'обычный (идемпотентный)'}")
    log_line("")
    log_line(f"Общий каталог запчастей z4 (уникальных): {total['parts_shared']}")
    log_line("")
    log_line(f"{'Таблица':<16}{'Из xlsm':>12}{'Вставлено':>12}{'В БД':>12}")
    log_line("-" * 52)
    log_line(f"{'works':<16}{total['works']:>12}{'':>12}{db_counts['works']:>12}")
    log_line(f"{'parts':<16}{total['parts_shared']:>12}{total['parts_inserted']:>12}"
             f"{db_counts['parts']:>12}")
    log_line(f"{'model_works':<16}{total['model_works']:>12}{'':>12}"
             f"{db_counts['model_works']:>12}")
    log_line(f"{'model_parts':<16}{total['model_parts']:>12}{'':>12}"
             f"{db_counts['model_parts']:>12}")
    log_line(f"{'matlib_entries':<16}{total['matlib_entries']:>12}{'':>12}"
             f"{db_counts['matlib_entries']:>12}")
    log_line(f"{'model_groups':<16}{'':>12}{'':>12}{db_counts['model_groups']:>12}")
    log_line("")
    log_line(f"user_version: {SCHEMA_VERSION}, journal_mode: WAL")
    dur = (datetime.now() - start_time).total_seconds()
    log_line(f"Время выполнения: {dur:.2f} с")
    log_line("=" * 60)

    return 0


def main() -> int:
    """Точка входа CLI."""
    parser = argparse.ArgumentParser(
        description="Конвертация base/models/*.xlsm в SysW.db (SQLite)"
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Полный пересозданный прогон (drop_tables=True)",
    )
    args = parser.parse_args()

    print("Запуск конвертера моделей -> SysW.db...")
    print("=" * 60)

    # Перезаписываем отчёт при каждом прогоне (чистый старт)
    with open(MIGRATION_REPORT_FILE, "w", encoding="utf-8") as f:
        f.write("")

    return run_migration(force=args.force)


if __name__ == "__main__":
    sys.exit(main())