#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
R-Z4GEN fallback-поиск ЗЧ через ReadGlobalPartByKey (план 1.1.3, п.5).

Назначение:
 1) Документирует структуру глобального справочника ЗЧ base/models/z4.xlsx
    (листы, колонки, заголовки) — те же источники, что читает
    Mod_ModelDB.ReadGlobalPartByKey.
 2) Воспроизводит логику поиска ReadGlobalPartByKey средствами Python:
    приоритет — № кат. (столбец B), fallback — наименование (столбец C),
    с нормализацией ключей (UPPER + trim) и числовым приведением.
 3) Прогоняет проверку на реальном z4.xlsx и выводит результат.

Ключи сопоставления (см. docs/table.md §8.2 и Mod_ModelDB.ReadGlobalPartByKey):
   B = Артикул (№ кат.)   -> ReadGlobalPartByKey(catNum, ...)
   C = Наименование       -> ReadGlobalPartByKey(..., partName)
   F = Цена, G = Кол-во ЗН.

Данные листа z4 читаются потоково (iter_rows) из-за большого объёма
(680 000+ строк): поиск выполняется по одному предзагруженному списку строк.
"""

from __future__ import annotations

import os
import sys

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl не установлен: pip install openpyxl") from exc

# Реальный источник, который использует GetGlobalPartsBasePath()
# (GetModelDBBasePath() = ThisWorkbook.Path & "\base\models\", Z4_NAME = "z4").
Z4_REL = os.path.join("base", "models", "z4.xlsx")
Z4_NAME = "z4"
DATA_START_ROW = 4  # Mod_Constants.DATA_START_ROW

# Столбцы листа z4, согласно BuildGlobalPart / ReadGlobalPartByKey.
COL_CATNUM = 1      # B (0-based индекс) — Артикул (№ кат.)
COL_NAME = 2        # C — Наименование
COL_PRICE = 5       # F — Цена
COL_QTY = 6         # G — Кол-во ЗН

HEADER_BY_COL = {
    1: "B=Артикул", 2: "C=Наименование", 5: "F=Цена", 6: "G=Кол-во ЗН",
}


def norm(value: object) -> str:
    """Нормализация ключа: UPPER + trim (как в ReadGlobalPartByKey)."""
    return str(value).strip().upper() if value is not None else ""


def build_part_from_row(row) -> dict:
    """Собирает словарь данных строки в формате BuildGlobalPart (A..H)."""
    get = lambda i: row[i] if i < len(row) else None  # noqa: E731
    return {
        "OutArticle": norm(get(COL_CATNUM)),
        "OutName": norm(get(COL_NAME)),
        "Price": float(get(COL_PRICE) or 0),
        "QtyZN": float(get(COL_QTY) or 0),
    }


def search(data_rows, key_cat: str, key_name: str):
    """
    Python-эквивалент Mod_ModelDB.ReadGlobalPartByKey по предзагруженным
    строкам (data_rows). Возвращает dict или None.
    """
    # 1. Приоритет — № кат. (B)
    if key_cat:
        for row in data_rows:
            if norm(row[COL_CATNUM]) == key_cat:
                return build_part_from_row(row)
    # 2. Fallback — наименование (C)
    if key_name:
        for row in data_rows:
            if norm(row[COL_NAME]) == key_name:
                return build_part_from_row(row)
    return None


def main() -> int:
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    path = os.path.join(root, Z4_REL)
    if not os.path.exists(path):
        print(f"[FAIL] Глобальный файл ЗЧ не найден: {path}")
        return 1

    print(f"Файл: {path} ({os.path.getsize(path):,} байт)")
    wb = load_workbook(path, read_only=True, data_only=True)
    print(f"Листы: {wb.sheetnames}")

    if Z4_NAME not in wb.sheetnames:
        print(f"[FAIL] Ожидаемый лист '{Z4_NAME}' отсутствует. Листы: {wb.sheetnames}")
        wb.close()
        return 1

    ws = wb[Z4_NAME]

    # Заголовки (строки 1..DATA_START_ROW-1) и примеры первых строк.
    header_rows = list(ws.iter_rows(min_row=1, max_row=DATA_START_ROW - 1, max_col=12, values_only=True))
    print("Заголовки (строки 1..3, первые 12 колонок):")
    for r, hr in enumerate(header_rows, start=1):
        print(f"  строка {r}: {list(hr)}")

    print("Примеры строк данных (первые 5):")
    sample = ws.iter_rows(min_row=DATA_START_ROW, max_row=DATA_START_ROW + 4, max_col=8, values_only=True)
    for r, srow in enumerate(sample, start=DATA_START_ROW):
        b = srow[COL_CATNUM] if COL_CATNUM < len(srow) else None
        c = srow[COL_NAME] if COL_NAME < len(srow) else None
        f = srow[COL_PRICE] if COL_PRICE < len(srow) else None
        g = srow[COL_QTY] if COL_QTY < len(srow) else None
        print(f"  строка {r}: B={b!r} C={c!r} F={f!r} G={g!r}")

    # Полная потоковая загрузка данных (одна строка = кортеж значений).
    data_rows = list(ws.iter_rows(min_row=DATA_START_ROW, values_only=True))
    wb.close()
    print(f"Загружено строк данных: {len(data_rows)}")

    ok = True

    # Пустые ключи -> Nothing (None).
    if search(data_rows, "", "") is not None:
        ok = False
        print("[FAIL] Пустые ключи должны давать Nothing (None).")
    else:
        print("[OK] Пустые ключи -> Nothing (None).")

    # Отсутствующий артикул -> Nothing (None).
    if search(data_rows, "__NO_SUCH_PART__", "") is not None:
        ok = False
        print("[FAIL] Отсутствующий артикул должен давать Nothing (None).")
    else:
        print("[OK] Отсутствующий артикул '__NO_SUCH_PART__' -> Nothing (None).")

    # Поиск по реальной первой строке данных (приоритет — № кат.).
    first = build_part_from_row(data_rows[0]) if data_rows else {"OutArticle": "", "OutName": ""}
    first_cat = first["OutArticle"]
    first_name = first["OutName"]
    hit = None
    if first_cat:
        hit = search(data_rows, first_cat, "")
    if hit is None and first_name:
        hit = search(data_rows, "", first_name)
    if hit is None:
        ok = False
        print(f"[FAIL] Не удалось найти первую строку данных по ключу "
              f"cat={first_cat!r} name={first_name!r}")
    else:
        print(f"[OK] Реальный запрос: cat={first_cat!r} name={first_name!r} -> "
              f"OutArticle={hit['OutArticle']!r} OutName={hit['OutName']!r} "
              f"Price={hit['Price']} QtyZN={hit['QtyZN']}")

    # Нормализация (регистр/пробелы).
    if first_cat:
        dirty = "  " + first_cat.lower() + "  "
        hit2 = search(data_rows, dirty, "")
        if hit2 is None:
            ok = False
            print(f"[FAIL] Нормализация ключа (lower+spaces) не сработала для {dirty!r}")
        else:
            print(f"[OK] Нормализация ключа {dirty!r} -> найдено (OutArticle={hit2['OutArticle']!r}).")

    # Числовой артикул: если ячейка B хранится как число, убеждаемся,
    # что поиск по строковому представлению числа работает.
    numeric_hits = 0
    for row in data_rows:
        val = row[COL_CATNUM] if COL_CATNUM < len(row) else None
        if isinstance(val, (int, float)):
            numeric_hits += 1
            num_hit = search(data_rows, norm(val), "")
            status = "OK" if num_hit is not None else "FAIL"
            if num_hit is None:
                ok = False
            print(f"[{status}] Числовой артикул {val!r} -> норм '{norm(val)}' "
                  f"найден={num_hit is not None}")
            break
    if numeric_hits == 0:
        print("[INFO] В файле не найдено числовых артикулов (проверка пропущена).")

    print("\nИТОГ:", "OK — источник z4.xlsx и логика fallback-поиска корректны."
          if ok else "ЕСТЬ ПРОБЛЕМЫ")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())