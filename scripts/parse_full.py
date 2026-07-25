# -*- coding: utf-8 -*-
"""
Полный парсинг ВСЕХ столбцов листов UAZw и UAZz4 из UAZ.xlsm
Использует openpyxl (без COM)
"""
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

# Путь к директории скрипта (для построения относительных путей)
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_DIR = os.path.normpath(os.path.join(_SCRIPT_DIR, '..'))


def col_letter(n: int) -> str:
    """Преобразует номер колонки (1-based) в буквенное обозначение (A, B, ..., Z, AA, ...)."""
    result = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def get_last_filled_col(ws, header_row: int = 3) -> int:
    """Определяет последнюю заполненную колонку по строке заголовков."""
    last = 1
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val is not None and str(val).strip() != '':
            last = c
    return last


def format_row(ws, row: int, last_col: int) -> str:
    """Форматирует строку как 'A=val | B=val | ...'."""
    parts = []
    for c in range(1, last_col + 1):
        letter = col_letter(c)
        val = ws.cell(row=row, column=c).value
        val_str = str(val) if val is not None else ''
        parts.append(f'{letter}={val_str}')
    return ' | '.join(parts)


def dump_sheet(wb, sheet_name: str, output: list, max_search_rows: int = 2000):
    """Дамп одного листа: заголовки, первые 5 строк, 3 примера с заполненными I и J."""
    output.append('========================================')
    output.append(f'ЛИСТ: {sheet_name}')
    output.append('========================================')

    ws = wb[sheet_name]
    print(f'Sheet {sheet_name}: max_col={ws.max_column}, max_row={ws.max_row}', flush=True)

    last_col = get_last_filled_col(ws)
    output.append(f'Последняя заполненная колонка (по строке 3): {last_col}')
    output.append('')

    # --- ЗАГОЛОВКИ ---
    output.append('--- ЗАГОЛОВКИ ВСЕХ КОЛОНОК (строка 3) ---')
    for c in range(1, last_col + 1):
        letter = col_letter(c)
        val = ws.cell(row=3, column=c).value
        val_str = str(val) if val is not None else ''
        output.append(f'{letter} ({c}): {val_str}')
    output.append('')

    # --- ПЕРВЫЕ 5 СТРОК ---
    output.append('--- ПЕРВЫЕ 5 СТРОК ДАННЫХ (4-8) ---')
    for r in range(4, min(9, ws.max_row + 1)):
        output.append(f'Строка {r}: {format_row(ws, r, last_col)}')
    output.append('')

    # --- 3 ПРИМЕРА СТРОК С ЗАПОЛНЕННЫМИ I и J ---
    output.append('--- 3 ПРИМЕРА СТРОК С ЗАПОЛНЕННЫМИ I и J ---')
    found = 0
    # Ограничиваем поиск первыми max_search_rows строками данных
    max_row = min(ws.max_row + 1, 4 + max_search_rows)
    for r in range(4, max_row):
        if found >= 3:
            break
        val_i = ws.cell(row=r, column=9).value
        val_j = ws.cell(row=r, column=10).value
        if val_i is not None and val_j is not None and str(val_i).strip() != '' and str(val_j).strip() != '':
            output.append(f'Строка {r}: {format_row(ws, r, last_col)}')
            found += 1

    if found < 3:
        output.append(f'[Найдено только {found} строк(и) с заполненными I и J в первых {max_search_rows} строках]')
    output.append('')


def main():
    # --- Определяем пути ---
    xlsm_path = os.path.join(_PROJECT_DIR, 'base', 'models', 'UAZ.xlsm')
    out_path = os.path.join(_PROJECT_DIR, 'plans', 'parsed_data_full.md')

    # --- Загружаем workbook ---
    print('Loading workbook...', flush=True)
    try:
        wb = load_workbook(xlsm_path, data_only=True)
    except FileNotFoundError:
        print(f'ERROR: Файл не найден: {xlsm_path}', flush=True)
        sys.exit(1)
    except Exception as e:
        print(f'ERROR: Не удалось загрузить workbook: {e}', flush=True)
        sys.exit(1)
    print('Workbook loaded.', flush=True)

    output = []

    # --- Обрабатываем листы ---
    for sheet_name in ('UAZw', 'UAZz4'):
        if sheet_name not in wb.sheetnames:
            print(f'WARNING: Лист "{sheet_name}" не найден в книге. Пропускаем.', flush=True)
            output.append(f'========================================')
            output.append(f'ЛИСТ: {sheet_name} — НЕ НАЙДЕН')
            output.append(f'========================================')
            output.append('')
            continue
        dump_sheet(wb, sheet_name, output)

    wb.close()

    # --- Сохраняем результат ---
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output))

    print(f'Done. Saved to {out_path}', flush=True)


if __name__ == '__main__':
    main()