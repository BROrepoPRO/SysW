#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Общая логика защиты листов шаблонов каталога base/templates/.

Единый источник поведения для build_templates.py и apply_protection_templates.py
(унификация v1.0.4). Согласованное правило:
  - UserInterfaceOnly=False — защита сохраняется в файле (реально пишется
    <sheetProtection> в XML), макросы работают через явные AllowEditRanges;
  - явные AllowEditRanges для зон ввода/данных (не надежда на UserInterfaceOnly);
  - FreezePanes на A4 (закреплены строки 1-3) на листах с заголовками.

Зоны (docs/table.md, раздел 0.4 / решение пользователя v1.0.4):
  work.xlsm, лист main:
      AllowEditRanges/разблокировка: B4, B5:B17, C1, Z1, данные D4:AB2000
      (доп. блокировка столбцов A:B);
  work.xlsm, spisok/libname: A2:J5000;
  work.xlsm, models:          A3:F5000;
  model.xlsm: A4:U2000 на всех листах + C1 на {GroupName} и z4;
  report0.xlsx: A2:Z5000.

Функции:
  apply_protection(ws, sheet_name, ...) — Protect + AllowEditRanges + FreezePanes;
  apply_freeze_only(ws)                  — только FreezePanes A4 (без защиты),
                                           для пустых work0/model0;
  verify_sheet_protection(path)          — проверка XML книги.
"""
import re
import zipfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --- Зоны ввода/данных по шаблонам ---
WORK_MAIN_EDIT = ["B4", "B5:B17", "C1", "Z1"]
WORK_MAIN_DATA = "D4:AB2000"     # данные работ/запчастей листа main
WORK_LIST_EDIT = "A2:J5000"      # spisok, libname
WORK_MODELS_EDIT = "A3:F5000"    # models

MODEL_DATA = "A4:U2000"          # все листы модельного шаблона
MODEL_EDIT_EXTRA = ["C1"]        # доп. зона на {GroupName} и z4

REPORT_EDIT = "A2:Z5000"         # report0.xlsx

# Заголовок AllowEditRanges
EDIT_RANGE_TITLE = "Ввод"


def _set_freeze_a4(ws):
    """Закрепляет строки 1-3 (FreezePanes на A4) через COM-окно Excel.

    Использует SplitRow=3 (закреплены 3 верхние строки) + FreezePanes=True.
    Для невидимой книги окно активируется принудительно.
    """
    try:
        ws.Activate()
        win = ws.Application.ActiveWindow
        # Сброс закрепления, затем фиксация 3 верхних строк (A4)
        win.SplitRow = 0
        win.SplitColumn = 0
        win.FreezePanes = False
        win.SplitRow = 3
        win.FreezePanes = True
    except Exception:
        # Закрепление через окно не всегда применимо к невидимой книге —
        # гарантированно устанавливается позднее через openpyxl на XML-уровне
        # (см. ensure_freeze_panes_after_save).
        pass


def _unlock_used(ws):
    """Разблокирует использованный диапазон (сброс предыдущих Locked=True)."""
    try:
        used = ws.UsedRange
        if used is not None:
            used.Locked = False
    except Exception:
        pass


def _resolve_zones(sheet_name, is_main, is_model, is_report):
    """Возвращает (edit_ranges) по типу шаблона и листа."""
    if is_report:
        # report0.xlsx: все листы (report, spisok)
        return [REPORT_EDIT]
    if is_model:
        # model.xlsm: A4:U2000 на всех листах + C1 на {GroupName} и z4
        if sheet_name in ("{GroupName}", "z4"):
            return MODEL_EDIT_EXTRA + [MODEL_DATA]
        return [MODEL_DATA]
    if is_main:
        # work.xlsm, лист main
        return [*WORK_MAIN_EDIT, WORK_MAIN_DATA]
    if sheet_name in ("spisok", "libname"):
        return [WORK_LIST_EDIT]
    if sheet_name == "models":
        return [WORK_MODELS_EDIT]
    # Прочие листы work-шаблона (например, _SETTINGS, если присутствует)
    return ["A4:AB2000"]


def _apply_protect(ws):
    """Включает защиту листа с UserInterfaceOnly=False (сохраняется в файле)."""
    try:
        ws.Protect(Password="", DrawingObjects=False, Contents=True,
                   Scenarios=False, AllowFormattingCells=False,
                   AllowFormattingColumns=False, AllowFormattingRows=False)
    except Exception as e:
        print(f"    [!] Protect на листе '{ws.Name}': {e}")


def _apply_allow_edit_ranges(ws, edit_ranges):
    """Добавляет явные AllowEditRanges для зон ввода/данных."""
    if not edit_ranges:
        return
    try:
        # Удаляем ранее существовавшие диапазоны
        for ar in list(ws.Protection.AllowEditRanges):
            ar.Delete()
        for rng in edit_ranges:
            ws.Protection.AllowEditRanges.Add(
                Title=EDIT_RANGE_TITLE, Range=ws.Range(rng))
    except Exception as e:
        print(f"    [!] AllowEditRanges на листе '{ws.Name}': {e}")


def apply_protection(ws, sheet_name, is_main=False, is_model=False, is_report=False):
    """Применяет Protect + AllowEditRanges + FreezePanes к листу ws.

    is_main   — лист main (доп. блокировка столбцов A:B).
    is_model  — листы модельного шаблона.
    is_report — листы шаблона отчёта.
    """
    # Снять защиту, если была
    try:
        ws.Unprotect()
    except Exception:
        pass

    # Сброс блокировок, затем блокировка строк 1-3
    _unlock_used(ws)
    ws.Rows("1:3").Locked = True

    # Зоны ввода/данных
    edit_ranges = _resolve_zones(sheet_name, is_main, is_model, is_report)

    # Разблокируем зоны ввода/данных
    for rng in edit_ranges:
        try:
            ws.Range(rng).Locked = False
        except Exception:
            pass

    # Для main: дополнительно блокируем столбцы A:B (кроме B4/B5:B17)
    if is_main:
        ws.Range("A:B").Locked = True
        for rng in ["B4", "B5:B17"]:
            try:
                ws.Range(rng).Locked = False
            except Exception:
                pass

    # FreezePanes на A4 (закреплены строки 1-3)
    _set_freeze_a4(ws)

    # Protect (UserInterfaceOnly=False — защита сохраняется в файле)
    _apply_protect(ws)

    # Явные AllowEditRanges
    _apply_allow_edit_ranges(ws, edit_ranges)


def apply_freeze_only(ws):
    """Только FreezePanes A4 без защиты — для пустых шаблонов work0/model0."""
    try:
        ws.Unprotect()
    except Exception:
        pass
    _set_freeze_a4(ws)


# ---------------------------------------------------------------------------
# Проверка итогового XML (без Excel, через zip + стандартную библиотеку)
# ---------------------------------------------------------------------------

_SHEET_XML_RE = re.compile(r'^xl/worksheets/sheet\d+\.xml$')
_PANE_RE = re.compile(r'<pane[^>]*xSplit="0"[^>]*ySplit="(\d+)"[^>]*/>')


def _extract_sheet_xmls(path):
    """Возвращает список (sheet_name, xml_text) из xlsx/xlsm-архива."""
    wb = load_workbook(str(path), keep_vba=True)
    # Порядок листов и их внутренние имена
    names = [ws.title for ws in wb.worksheets]

    results = []
    with zipfile.ZipFile(str(path)) as zf:
        # Сортируем листы по номеру в имени файла
        sheet_files = sorted(
            (n for n in zf.namelist() if _SHEET_XML_RE.match(n)),
            key=lambda n: int(re.search(r'sheet(\d+)\.xml$', n).group(1)),
        )
        for i, sheet_file in enumerate(sheet_files):
            xml_text = zf.read(sheet_file).decode("utf-8")
            name = names[i] if i < len(names) else sheet_file
            results.append((name, xml_text))
    return results


def ensure_freeze_panes_after_save(path):
    """Гарантированно закрепляет строки 1-3 (FreezePanes A4) на XML-уровне.

    Вызывается ПОСЛЕ сохранения книги через Excel COM. Открывает .xlsm/.xlsx
    через openpyxl (keep_vba=True) и задаёт freeze_panes = 'A4' каждому листу.
    Важно: openpyxl переписывает xl/worksheets/*.xml и часть архивных записей,
    но сохраняет vbaProject.bin для .xlsm (keep_vba=True). Используется как
    страховка, когда COM-окно не смогло установить закрепление для невидимой
    книги. При необходимости шрифт/формат сохраняются; допустимо запускать
    повторно (идемпотентно).
    """
    is_xlsm = str(path).lower().endswith(".xlsm")
    wb = load_workbook(str(path), keep_vba=is_xlsm)
    changed = False
    for ws in wb.worksheets:
        if ws.freeze_panes != "A4":
            ws.freeze_panes = "A4"
            changed = True
    if changed:
        wb.save(str(path))
    return changed


def _all_sheets_frozen_a4(path):
    """True, если ВСЕ листы книги уже имеют FreezePanes A4 (frozen ySplit=3).

    Анализ ведётся напрямую по zip-архиву (без openpyxl): проверяется, что
    в каждом xl/worksheets/sheetN.xml присутствует узел <pane> с xSplit="0"
    и ySplit="3". Используется для идемпотентности: если файл уже корректен,
    повторная правка не выполняется (не переписываем архив без необходимости).
    """
    with zipfile.ZipFile(str(path)) as zf:
        sheet_files = [n for n in zf.namelist() if _SHEET_XML_RE.match(n)]
        if not sheet_files:
            return False
        for n in sheet_files:
            xml_text = zf.read(n).decode("utf-8")
            pm = re.search(_PANE_RE, xml_text)
            if not pm or int(pm.group(1)) != 3:
                return False
    return True


def apply_freeze_panes_to_models(models_dir):
    """Устанавливает FreezePanes A4 (закреплены строки 1-3) всем листам
    модельных файлов base/models/*.xlsm НА XML-УРОВНЕ (ВАРИАНТ C, v1.0.9).

    ИСПРАВЛЕНИЕ v1.0.9: прежний метод через openpyxl (keep_vba=True) признан
    опасным — валидация показала, что пересохранение модельных .xlsm через
    openpyxl делает файлы нечитаемыми для Excel COM (после сохранения
    Workbooks.Open возвращает ошибку -2147352567). Поэтому применяется
    точечная правка ТОЛЬКО узла <pane> внутри zip-архива (apply_freeze_panes_xml):
    остальные записи (включая sheet*.xml и vbaProject.bin, если он есть)
    переносятся байт-в-байт без изменений. Модельные файлы не содержат
    VBA-проекта (vbaProject.bin отсутствует), но принцип максимального
    сохранения исходного формата и целостности критических файлов ([E3])
    соблюдён полностью.

    Функция идемпотентна: если во всех листах уже frozen A4 — файл не трогается.
    """
    changed_any = False
    for path in sorted(models_dir.glob("*.xlsm")):
        try:
            if _all_sheets_frozen_a4(path):
                continue
        except Exception as exc:
            print(f"  [!] Не удалось прочитать pane модели {path.name}: {exc}")
            continue
        apply_freeze_panes_xml(path)
        changed_any = True
        print(f"  FreezePanes A4 применён к {path.name} (XML-уровень)")
    return changed_any


def verify_sheet_protection(path):
    """Проверяет XML книги: <sheetProtection>, <allowEditRanges>, <pane> frozen A4.

    Не требует lxml — используется только стандартная библиотека (zipfile, re)
    и openpyxl для получения имён листов. Возвращает список строк-отчётов вида:
        main | protect=да | allowEdit=да (B4, ...) | pane=A4
    """
    reports = []
    for name, xml_text in _extract_sheet_xmls(path):
        prot = "да" if "<sheetProtection" in xml_text else "нет"

        # AllowEditRanges: ищем блок <allowEditRanges>...</allowEditRanges>
        aem = re.search(r'<allowEditRanges[^>]*>(.*?)</allowEditRanges>',
                        xml_text, re.S)
        if aem and re.search(r'<rangeEdit[^>]*>', aem.group(1)):
            sqrefs = re.findall(r'sqref="([^"]+)"', aem.group(1))
            allow_edit = "да (" + ", ".join(sqrefs) + ")"
        else:
            allow_edit = "нет"

        # FreezePanes: pane с ySplit (строки закрепления); frozen A4 => ySplit=3
        pane = "нет"
        pm = re.search(_PANE_RE, xml_text)
        if pm:
            ysplit = int(pm.group(1))
            pane = "A4" if ysplit == 3 else f"frozen ySplit={ysplit}"
        else:
            if "<pane" in xml_text:
                pane = "присутствует (не frozen A4)"

        reports.append(f"{name} | protect={prot} | allowEdit={allow_edit} | pane={pane}")
    return reports


def list_allowed_edits(path):
    """Краткий список зон AllowEditRanges по листам (для резюме)."""
    out = []
    for name, xml_text in _extract_sheet_xmls(path):
        aem = re.search(r'<allowEditRanges[^>]*>(.*?)</allowEditRanges>',
                        xml_text, re.S)
        sqrefs = re.findall(r'sqref="([^"]+)"', aem.group(1)) if aem else []
        out.append(f"{name}: {sqrefs}")
    return out


# ---------------------------------------------------------------------------
# ВАРИАНТ D: защита на уровне XML (точечная правка sheet*.xml в zip-архиве,
# без пересохранения книги через openpyxl). Максимально сохраняет исходный
# формат: изменяются только узлы <pane>, <sheetProtection>, <allowEditRanges>,
# остальные записи архива переносятся без изменений.
#
# Применяется ПОСЛЕ сохранения книги Excel COM (Locked-флаги ячеек уже
# выставлены COM-ом, формат сохранён Excel-ом). Здесь добавляется то, что
# COM не смог надёжно записать: allowEditRanges (ошибка -2147352567 из-за
# предварительно разблокированных зон) и FreezePanes A4 (pane).
# ---------------------------------------------------------------------------
import xml.etree.ElementTree as _ET  # локальный импорт (только stdlib)
import shutil as _shutil

# FreezePanes A4: закреплены строки 1-3 (ySplit=3), столбцы не отрываются
_SHEET_PANE_FROZEN_A4 = ('<pane xSplit="0" ySplit="3" topLeftCell="A4" '
                         'activePane="bottomLeft" state="frozen"/>')
_SHEET_SELECTION_BL = ('<selection pane="bottomLeft" activeCell="A4" '
                       'sqref="A4"/>')
# Защита листа: запрет изменений содержимого/объектов/сценариев,
# НО разрешён автофильтр (enableAutoFilter) — нужен макросам поиска/фильтрации.
_SHEET_PROTECTION = ('<sheetProtection sheet="1" objects="1" scenarios="1" '
                     'autoFilter="1" '
                     'selectLockedCells="1" selectUnlockedCells="1"/>')
_EDIT_RANGE_NAME = "Ввод"


def _sheet_name_map(path):
    """Возвращает {путь в архиве 'xl/worksheets/sheetN.xml': имя листа}."""
    mapping = {}
    try:
        with zipfile.ZipFile(str(path)) as zf:
            wb = zf.read('xl/workbook.xml').decode('utf-8')
            rels = zf.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        # rId -> Target (worksheets/sheetN.xml) — атрибуты в любом порядке,
        # допускается префикс '../' или абсолютный '/xl/...'
        rid2tgt = {}
        for rel in re.finditer(r'<Relationship\b[^>]*/?>', rels):
            el = rel.group(0)
            if 'worksheets/sheet' not in el:
                continue
            mid = re.search(r'Id="(rId\d+)"', el)
            mtgt = re.search(r'Target="([^"]*worksheets/sheet\d+\.xml)"', el)
            if mid and mtgt:
                tgt = mtgt.group(1).replace('../', '').lstrip('/')
                if not tgt.startswith('xl/'):
                    tgt = 'xl/' + tgt
                rid2tgt[mid.group(1)] = tgt
        # sheet -> rId (атрибуты в любом порядке)
        for s in re.finditer(r'<sheet\b[^>]*/?>', wb):
            el = s.group(0)
            mname = re.search(r'name="([^"]+)"', el)
            mrid = re.search(r'r:id="(rId\d+)"', el)
            if mname and mrid:
                tgt = rid2tgt.get(mrid.group(1))
                if tgt:
                    mapping[tgt] = mname.group(1)
    except Exception:
        pass
    return mapping


def _inject_sheet_xml(xml_text, sheet_name, zones):
    """Встраивает pane/защиту/allowEditRanges в текст sheet*.xml.

    zones — список диапазонов (str) для allowEditRanges либо None (не создавать).
    """
    # 1) FreezePanes A4 в первый <sheetView>
    if '<pane' in xml_text:
        # Заменяем первый pane на frozen A4 (count=1)
        xml_text = re.sub(r'<pane[^>]*/>', _SHEET_PANE_FROZEN_A4,
                          xml_text, count=1)
    else:
        mv = re.search(r'(<sheetView\b[^>]*>)(.*?)(</sheetView>)',
                       xml_text, re.S)
        if mv and '<pane' not in mv.group(2):
            inner = _SHEET_PANE_FROZEN_A4 + _SHEET_SELECTION_BL + mv.group(2)
            xml_text = (xml_text[:mv.start()] + mv.group(1) + inner +
                        mv.group(3) + xml_text[mv.end():])

    # 2) Удаляем старые узлы sheetProtection и allowEditRanges (если были)
    xml_text = re.sub(r'<sheetProtection[^>]*/>', '', xml_text)
    xml_text = re.sub(r'<allowEditRanges>.*?</allowEditRanges>', '',
                      xml_text, flags=re.S)

    # 3) Собираем блок защиты + AllowEditRanges (если заданы зоны)
    # Порядок узлов по схеме CT_Worksheet:
    #   sheetData -> sheetCalcPr -> sheetProtection -> protectedRanges
    #   (protectedRanges ДОЛЖНЫ идти после sheetProtection и ДО autoFilter).
    block = _SHEET_PROTECTION
    if zones:
        sqref = ' '.join(zones)
        block += ('<allowEditRanges><rangeEdit name="%s" sqref="%s"/>'
                  '</allowEditRanges>' % (_EDIT_RANGE_NAME, sqref))

    # Вставляем блок ПОСЛЕ закрывающего </sheetData> (это правильная позиция:
    # вставка перед <sheetData> делала книгу нечитаемой для Excel — 0x800A03EC).
    anchor = re.search(r'</sheetData>', xml_text)
    if anchor is not None:
        insert_at = anchor.end()
        # Если сразу после sheetData следует sheetCalcPr — защита должна быть
        # ПОСЛЕ него (sheetCalcPr предшествует sheetProtection по схеме).
        mcalc = re.match(
            r'\s*(?:<sheetCalcPr[^>]*/>|<sheetCalcPr\b.*?</sheetCalcPr>)',
            xml_text[insert_at:], re.S)
        if mcalc:
            insert_at += mcalc.end()
        xml_text = (xml_text[:insert_at] + block + xml_text[insert_at:])
    else:
        # Резерв: нет sheetData (аномальный лист) — перед закрытием </worksheet>
        wend = xml_text.rfind('</worksheet>')
        if wend != -1:
            xml_text = xml_text[:wend] + block + xml_text[wend:]
    return xml_text


def apply_protection_xml(path, zone_map):
    """Применяет защиту на уровне XML к книге.

    zone_map: {имя листа: [диапазоны для allowEditRanges] | None}.
    Переписывает zip: правит только sheet*.xml, остальное без изменений.
    """
    tmp = str(path) + '.prot.tmp'
    name_map = _sheet_name_map(path)
    with zipfile.ZipFile(str(path), 'r') as zin, \
            zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if _SHEET_XML_RE.match(item.filename):
                sheet_name = name_map.get(item.filename)
                if sheet_name is not None:
                    zones = zone_map.get(sheet_name)
                    if zones is None and '__default__' in zone_map:
                        zones = zone_map['__default__']
                    text = data.decode('utf-8')
                    text = _inject_sheet_xml(text, sheet_name, zones)
                    data = text.encode('utf-8')
            zout.writestr(item, data)
    _shutil.move(tmp, str(path))


def build_zone_map(template_type):
    """Возвращает zone_map для apply_protection_xml по типу шаблона."""
    z = {}
    if template_type == 'work':
        z['main'] = [*WORK_MAIN_EDIT, WORK_MAIN_DATA]          # B4,B5:B17,C1,Z1,D4:AB2000
        z['spisok'] = [WORK_LIST_EDIT]                          # A2:J5000
        z['libname'] = [WORK_LIST_EDIT]                         # A2:J5000
        z['models'] = [WORK_MODELS_EDIT]                        # A3:F5000
        # прочие листы (например, _SETTINGS) — широкая зона
        z['_SETTINGS'] = ["A4:AB2000"]
    elif template_type == 'model':
        for nm in ("{GroupName}", "z4"):
            z[nm] = MODEL_EDIT_EXTRA + [MODEL_DATA]            # C1 + A4:U2000
        z.setdefault('__default__', [MODEL_DATA])               # прочие листы
    elif template_type == 'report':
        z['__default__'] = [REPORT_EDIT]                        # A2:Z5000
    return z


def apply_freeze_panes_xml(path):
    """Только FreezePanes A4 на XML-уровне (без защиты) — для пустых шаблонов
    work0/model0. Правит только <pane> в sheet*.xml, остальное не меняет."""
    tmp = str(path) + '.fr.tmp'
    with zipfile.ZipFile(str(path), 'r') as zin, \
            zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if _SHEET_XML_RE.match(item.filename):
                text = data.decode('utf-8')
                if '<pane' in text:
                    text = re.sub(r'<pane[^>]*/>', _SHEET_PANE_FROZEN_A4,
                                  text, count=1)
                else:
                    mv = re.search(r'(<sheetView\b[^>]*>)(.*?)(</sheetView>)',
                                   text, re.S)
                    if mv and '<pane' not in mv.group(2):
                        inner = (_SHEET_PANE_FROZEN_A4 + _SHEET_SELECTION_BL +
                                 mv.group(2))
                        text = (text[:mv.start()] + mv.group(1) + inner +
                                mv.group(3) + text[mv.end():])
                data = text.encode('utf-8')
            zout.writestr(item, data)
    _shutil.move(tmp, str(path))